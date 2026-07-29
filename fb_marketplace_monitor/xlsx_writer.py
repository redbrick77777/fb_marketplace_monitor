"""Local XLSX writer - an alternative to Google Sheets that needs no cloud
setup at all. Creates the workbook/worksheet if they don't exist yet, and
appends new rows on each run.

Each call opens, edits, and saves the file fresh rather than holding any
long-lived state - simple, and safe for a single cron run at a time. If the
file happens to be open in Excel/Numbers when cron fires, the OS will block
the save and we raise a clear error rather than silently losing data (the
caller is responsible for not marking those listings as "seen" until the
write actually succeeds - see monitor.py).
"""
from __future__ import annotations

import logging
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

HEADER_ROW = ["Listing ID", "Title", "Price", "Link", "Date Found", "Watch"]


class XlsxWriter:
    def append_rows(self, file_path: str, tab_name: str, rows: list[list]) -> None:
        if not rows:
            return

        path = Path(file_path)
        resolved = path.resolve()  # absolute path, symlinks/.. resolved - no ambiguity
        logger.info(
            "xlsx_writer: about to write %d row(s) to '%s' (resolved: %s), "
            "current working directory is '%s'",
            len(rows), file_path, resolved, os.getcwd(),
        )

        path.parent.mkdir(parents=True, exist_ok=True)

        if path.exists():
            try:
                workbook = load_workbook(path)
            except Exception as exc:
                raise RuntimeError(
                    f"Couldn't open '{file_path}' - is it corrupted, or not a "
                    f"valid .xlsx file? ({exc})"
                ) from exc
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)  # drop the default blank "Sheet"

        if tab_name in workbook.sheetnames:
            worksheet = workbook[tab_name]
            if worksheet.max_row == 1 and all(c.value is None for c in worksheet[1]):
                worksheet.append(HEADER_ROW)
        else:
            worksheet = workbook.create_sheet(title=tab_name)
            worksheet.append(HEADER_ROW)

        for row in rows:
            worksheet.append(row)

        try:
            workbook.save(path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Couldn't save '{file_path}' - it's probably open in Excel/Numbers "
                "right now. Close it and this will retry on the next scheduled run "
                "(nothing has been lost - these listings will be picked up again)."
            ) from exc

        # Self-check: verify the file is actually there, in THIS process,
        # right after save() returned - rather than trusting that a clean
        # return from openpyxl necessarily means a file landed on disk where
        # we think it did (e.g. permission quirks, unusual mounts, etc.).
        if resolved.exists():
            size = resolved.stat().st_size
            logger.info(
                "xlsx_writer: confirmed '%s' exists on disk immediately after "
                "save (size=%d bytes, worksheet='%s', total rows now=%d)",
                resolved, size, tab_name, worksheet.max_row,
            )
        else:
            logger.error(
                "xlsx_writer: workbook.save() returned without error, but "
                "'%s' does NOT exist immediately afterward! This means "
                "something unusual is happening at the filesystem/openpyxl "
                "level, not in our own logic - please report this exact "
                "path and resolved path back for further debugging.",
                resolved,
            )
